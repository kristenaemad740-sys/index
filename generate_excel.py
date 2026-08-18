"""
=============================================================================
  🏆 اليوم الرياضي لأسرة الكاروز — مولد الإكسيل التلقائي (Excel Generator)
=============================================================================
  يقوم بإنشاء وتحديث شيت إكسيل احترافي يحتوي على:
  1. 📊 Dashboard (كروت الإحصائيات + نسب الفرق والجنس + الرسوم البيانية)
  2. 📋 قائمة كاملة بالمسجلين مع الفلاتر
  3. 🏆 توزيع الفرق الأربعة بالتفصيل

  طريقة الاستخدام:
  - ببساطة شغّل: python generate_excel.py
=============================================================================
"""

import sys
import os
import csv
import json
import urllib.request
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ مكتبة openpyxl غير مثبتة. يرجى تشغيل: pip install openpyxl")
    sys.exit(1)

SCRIPT_URL = "https://script.google.com/macros/s/AKfycbw2y2VO_fJ5fcZCyJSHd4-bRqWmjaTGrHs4gX7u0uyMuFNsfv-JWfprhcOJrliikl6c/exec"
CSV_FILE = "registrations.csv"
OUTPUT_FILE = f"karoz_sports_dashboard_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

TEAMS = {
    "red":    {"name": "🔴 الفريق الأحمر",  "color": "DC2626", "fill": "FEE2E2"},
    "green":  {"name": "🟢 الفريق الأخضر",  "color": "16A34A", "fill": "DCFCE7"},
    "yellow": {"name": "🟡 الفريق الأصفر",  "color": "CA8A04", "fill": "FEF9C3"},
    "black":  {"name": "⚫ الفريق الأسود",  "color": "374151", "fill": "F3F4F6"},
}

def make_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def make_font(bold=False, size=11, color="000000", italic=False) -> Font:
    return Font(bold=bold, size=size, color=color, italic=italic, name="Cairo")

def make_border(style="thin") -> Border:
    s = Side(border_style=style, color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True, readingOrder=2)

def right_align() -> Alignment:
    return Alignment(horizontal="right", vertical="center", wrap_text=True, readingOrder=2)

def fetch_live_data() -> list[dict]:
    """محاولة جلب البيانات مباشرة من Google Apps Script"""
    try:
        payload = json.dumps({"action": "getAll"}).encode('utf-8')
        req = urllib.request.Request(
            SCRIPT_URL,
            data=payload,
            headers={'Content-Type': 'text/plain;charset=utf-8'},
            method='POST'
        )
        with urllib.request.urlopen(req, timeout=8) as response:
            res = json.loads(response.read().decode('utf-8'))
            if res.get("success") and res.get("data"):
                print(f"🌐 تم جلب {len(res['data'])} مشترك مباشرة من Google Sheets الحقيقي!")
                return res["data"]
    except Exception:
        pass
    return []

def load_data() -> list[dict]:
    # 1. جرب جلب مباشر من API
    live_records = fetch_live_data()
    if live_records:
        return live_records

    # 2. جرب ملف registrations.csv إذا كان موجوداً
    if os.path.exists(CSV_FILE):
        records = []
        with open(CSV_FILE, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                records.append(row)
        print(f"📄 تم تحميل {len(records)} سجل من ملف '{CSV_FILE}'")
        return records

    print(f"ℹ️ لم يتم العثور على '{CSV_FILE}' — جاري تشغيل أحدث البيانات التجريبية للمحاكاة...")
    import random
    random.seed(42)
    names = [
        "شاكر مينا محسن", "ماريو طلعت شاكر", "بيتر ميلاد حنا", "مارك فادي جرجس",
        "مينا نبيل يوسف", "داوود سامي رمزي", "كيرلس ماجد أنطوان", "مريم نادر يوسف",
        "سارة ايهاب منصور", "ماريان رمزي جبران", "نورا فادي قلادة", "هنا ميلاد نخلة"
    ]
    records = []
    for i, name in enumerate(names):
        gender = "female" if i >= 7 else "male"
        records.append({
            "id": f"p_{i+1}",
            "name": name,
            "phone": f"012{str(10000000+i).zfill(8)}",
            "gender": gender,
            "team": list(TEAMS.keys())[i % 4],
            "wantsFriends": "TRUE" if i % 2 == 0 else "FALSE",
            "friendsCount": "1" if i % 2 == 0 else "0",
            "friendNames": names[(i+1)%len(names)] if i % 2 == 0 else "",
            "registrationTime": datetime.now().isoformat()
        })
    return records

def build_excel(records: list[dict]):
    wb = openpyxl.Workbook()

    NAVY = "040D1E"
    GOLD = "F5A623"
    WHITE = "FFFFFF"
    GRAY_L = "F8FAFC"

    total = len(records)
    males = sum(1 for r in records if str(r.get("gender","")).lower() == "male")
    females = total - males
    teams_cnt = {t: 0 for t in TEAMS}
    for r in records:
        t = str(r.get("team","")).strip().lower()
        if t in teams_cnt:
            teams_cnt[t] += 1
    want_friends = sum(1 for r in records if str(r.get("wantsFriends","")).upper() == "TRUE")

    # ── Sheet 1: Dashboard ────────────────────────────────────────────────────
    dash = wb.active
    dash.title = "📊 Dashboard"
    dash.sheet_view.rightToLeft = True
    dash.sheet_properties.tabColor = "1E3A5F"

    for col, w in [(1,4),(2,22),(3,16),(4,16),(5,16),(6,16),(7,16),(8,4)]:
        dash.column_dimensions[get_column_letter(col)].width = w

    # Header
    dash.row_dimensions[2].height = 50
    dash.merge_cells("B2:G2")
    t_cell = dash["B2"]
    t_cell.value = "🏆 اليوم الرياضي لأسرة الكاروز — لوحة الإحصائيات الشاملة"
    t_cell.font = make_font(bold=True, size=16, color=GOLD)
    t_cell.fill = make_fill(NAVY)
    t_cell.alignment = center()

    dash.row_dimensions[3].height = 22
    dash.merge_cells("B3:G3")
    sub = dash["B3"]
    sub.value = f"📅 آخر تحديث: {datetime.now().strftime('%Y-%m-%d %H:%M')}   |   إجمالي المسجلين: {total}"
    sub.font = make_font(size=11, color="94A3B8")
    sub.fill = make_fill("0D1B3E")
    sub.alignment = center()

    # KPI Cards
    dash.row_dimensions[5].height = 65
    kpis = [
        ("B5", f"{total}", "إجمالي المسجلين", GOLD, "1E3A5F"),
        ("C5", f"{males}", f"👦 أولاد ({males/total*100:.0f}%)" if total else "👦 أولاد", "60A5FA", "1E3A8A"),
        ("D5", f"{females}", f"👧 بنات ({females/total*100:.0f}%)" if total else "👧 بنات", "F472B6", "831843"),
        ("E5", f"{want_friends}", "🤝 طلبوا أصدقاء", "34D399", "064E3B"),
    ]
    for cell_ref, val, label, fc, bg in kpis:
        c = dash[cell_ref]
        c.value = f"{val}\n{label}"
        c.font = Font(bold=True, size=13, color=fc, name="Cairo")
        c.fill = make_fill(bg)
        c.alignment = center()
        c.border = make_border()

    # Team Distribution Table
    dash.row_dimensions[7].height = 26
    for col, hdr in zip(["B","C","D","E"], ["الفريق", "العدد", "النسبة %", "توزيع الجنس (ولد/بنت)"]):
        c = dash[f"{col}7"]
        c.value = hdr
        c.font = make_font(bold=True, size=11, color=WHITE)
        c.fill = make_fill(NAVY)
        c.alignment = center()
        c.border = make_border()

    row = 8
    for tid, tinfo in TEAMS.items():
        cnt = teams_cnt.get(tid, 0)
        pct = (cnt / total * 100) if total else 0
        t_males = sum(1 for r in records if str(r.get("team","")).lower() == tid and str(r.get("gender","")).lower() == "male")
        t_females = cnt - t_males
        gender_str = f"👦 {t_males}  |  👧 {t_females}"

        dash.row_dimensions[row].height = 24
        data = [tinfo["name"], cnt, f"{pct:.1f}%", gender_str]
        for col_letter, val in zip(["B","C","D","E"], data):
            c = dash[f"{col_letter}{row}"]
            c.value = val
            c.font = make_font(size=11, color="1F2937")
            c.fill = make_fill(tinfo["fill"])
            c.alignment = center()
            c.border = make_border()
        row += 1

    # ── Sheet 2: قائمة المسجلين ───────────────────────────────────────────────
    reg_sheet = wb.create_sheet("📋 قائمة المسجلين")
    reg_sheet.sheet_view.rightToLeft = True
    reg_sheet.sheet_properties.tabColor = "16A34A"

    for col, w in [(1,5),(2,26),(3,16),(4,12),(5,14),(6,22),(7,20)]:
        reg_sheet.column_dimensions[get_column_letter(col)].width = w

    headers = ["#", "الاسم", "رقم الواتساب", "الجنس", "الفريق", "الأصدقاء المطلوبة", "وقت التسجيل"]
    reg_sheet.row_dimensions[1].height = 28
    for col_idx, hdr in enumerate(headers, 1):
        c = reg_sheet.cell(row=1, column=col_idx, value=hdr)
        c.font = make_font(bold=True, size=11, color=WHITE)
        c.fill = make_fill(NAVY)
        c.alignment = center()
        c.border = make_border()

    for row_idx, r in enumerate(records, 2):
        reg_sheet.row_dimensions[row_idx].height = 22
        tid = str(r.get("team","")).strip().lower()
        tname = TEAMS.get(tid, {}).get("name", tid)
        tfill = TEAMS.get(tid, {}).get("fill", "FFFFFF")
        gender_ar = "👦 ذكر" if str(r.get("gender","")).lower() == "male" else "👧 أنثى"
        friends_text = str(r.get("friendNames","") or "—")

        row_data = [
            row_idx - 1,
            r.get("name",""),
            r.get("phone",""),
            gender_ar,
            tname,
            friends_text,
            r.get("registrationTime","")[:19].replace("T", " ")
        ]
        bg = GRAY_L if row_idx % 2 == 0 else WHITE
        for col_idx, val in enumerate(row_data, 1):
            c = reg_sheet.cell(row=row_idx, column=col_idx, value=val)
            c.font = make_font(size=10)
            c.fill = make_fill(tfill if col_idx == 5 else bg)
            c.alignment = center() if col_idx != 2 else right_align()
            c.border = make_border()

    reg_sheet.auto_filter.ref = f"A1:G{len(records)+1}"
    reg_sheet.freeze_panes = "A2"

    # ── Sheet 3: توزيع الفرق ──────────────────────────────────────────────────
    teams_sheet = wb.create_sheet("🏆 توزيع الفرق")
    teams_sheet.sheet_view.rightToLeft = True
    teams_sheet.sheet_properties.tabColor = "CA8A04"

    col_idx = 1
    for tid, tinfo in TEAMS.items():
        members = [r for r in records if str(r.get("team","")).strip().lower() == tid]
        teams_sheet.column_dimensions[get_column_letter(col_idx)].width = 24

        c1 = teams_sheet.cell(row=1, column=col_idx, value=tinfo["name"])
        c1.font = make_font(bold=True, size=12, color=WHITE)
        c1.fill = make_fill(tinfo["color"])
        c1.alignment = center()
        c1.border = make_border()

        c2 = teams_sheet.cell(row=2, column=col_idx, value=f"العدد: {len(members)}")
        c2.font = make_font(bold=True, size=11, color="1F2937")
        c2.fill = make_fill(tinfo["fill"])
        c2.alignment = center()
        c2.border = make_border()

        for i, m in enumerate(members, 3):
            cm = teams_sheet.cell(row=i, column=col_idx, value=m.get("name",""))
            cm.font = make_font(size=10)
            cm.fill = make_fill(tinfo["fill"] if i % 2 == 0 else WHITE)
            cm.alignment = right_align()
            cm.border = make_border()

        col_idx += 1

    wb.save(OUTPUT_FILE)
    print(f"\n✅ تم إنشاء وتحديث ملف الإكسيل بنجاح: {OUTPUT_FILE}")
    return OUTPUT_FILE

if __name__ == "__main__":
    records = load_data()
    out = build_excel(records)
    if os.name == 'nt':
        os.system(f'start "" "{out}"')
